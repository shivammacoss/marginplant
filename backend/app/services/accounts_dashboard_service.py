"""Accounts Dashboard service — broker totals, per-user PNL breakdown, exports.

Reuses the PNL calculation pattern from pnl_sharing_service / admin_settlement_service
but operates independently (no PnlSharingAgreement required).
"""

from __future__ import annotations

import logging
from datetime import datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Any

from beanie import PydanticObjectId
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.models.pnl_sharing import AgreementStatus, AgreementType, PnlSharingAgreement
from app.models.position import Position, PositionStatus
from app.models.trade import Trade
from app.models.transaction import TransactionType, WalletTransaction
from app.models.user import User, UserRole, UserStatus
from app.models.wallet import Wallet
from app.services import market_data_service
from app.services.admin_settlement_service import _realised_inr
from app.utils.decimal_utils import quantize_money, to_decimal
from app.utils.time_utils import IST

logger = logging.getLogger(__name__)

TRADING_ROLES = [UserRole.CLIENT.value, UserRole.DEALER.value, UserRole.MASTER.value]


# ── Week options for dropdown ────────────────────────────────────────

def generate_week_options(num_weeks: int = 16) -> list[dict[str, str]]:
    """Last N IST weeks (Monday–Sunday). Most recent first."""
    today = datetime.now(IST).date()
    monday = today - timedelta(days=today.weekday())
    weeks = []
    for i in range(num_weeks):
        w_mon = monday - timedelta(weeks=i)
        w_sun = w_mon + timedelta(days=6)
        weeks.append({
            "label": f"Week_{w_mon.isoformat()}",
            "start": w_mon.isoformat(),
            "end": w_sun.isoformat(),
        })
    return weeks


# ── Broker client lookup ─────────────────────────────────────────────

async def _broker_client_ids(broker_id: PydanticObjectId) -> list[PydanticObjectId]:
    """Direct clients of a broker (assigned_broker_id == broker_id, trading roles)."""
    coll = User.get_motor_collection()
    cursor = coll.find(
        {
            "assigned_broker_id": broker_id,
            "role": {"$in": TRADING_ROLES},
            "status": {"$ne": UserStatus.CLOSED.value},
        },
        {"_id": 1},
    )
    return [doc["_id"] async for doc in cursor]


async def _entity_pool_ids(
    entity_id: PydanticObjectId,
    entity_role: str,
) -> list[PydanticObjectId]:
    """User IDs in an entity's pool based on role."""
    coll = User.get_motor_collection()
    query: dict[str, Any] = {
        "role": {"$in": TRADING_ROLES},
        "status": {"$ne": UserStatus.CLOSED.value},
    }
    if entity_role in (UserRole.BROKER.value, "BROKER"):
        query["assigned_broker_id"] = entity_id
    elif entity_role in (UserRole.ADMIN.value, "ADMIN"):
        query["assigned_admin_id"] = entity_id
    elif entity_role in (UserRole.SUPER_ADMIN.value, "SUPER_ADMIN"):
        pass  # all users
    else:
        query["assigned_broker_id"] = entity_id
    cursor = coll.find(query, {"_id": 1})
    return [doc["_id"] async for doc in cursor]


# ── Broker totals (PNL sharing snapshot) ─────────────────────────────

async def compute_broker_totals(
    entity_id: PydanticObjectId,
    start_utc: datetime | None,
    end_utc: datetime | None,
) -> dict[str, Any]:
    """Compute NET CLIENT PNL, BKG, TOTAL, SETTLEMENT, ACTUAL PNL, SHARING for a broker."""
    entity = await User.get(entity_id)
    if not entity:
        return _empty_broker_totals()

    user_ids = await _broker_client_ids(entity_id)

    net_client_pnl = Decimal("0")
    net_client_bkg = Decimal("0")
    total_deposits = Decimal("0")
    total_withdrawals = Decimal("0")

    if user_ids:
        fallback_usd_inr = to_decimal(market_data_service.get_usd_inr_rate())

        # Build date filter once
        date_filter: dict[str, Any] | None = None
        if start_utc or end_utc:
            date_filter = {}
            if start_utc:
                date_filter["$gte"] = start_utc
            if end_utc:
                date_filter["$lte"] = end_utc

        # Realized PNL — ALL positions (open with partial close PNL + closed)
        if date_filter:
            pos_query: dict[str, Any] = {
                "user_id": {"$in": user_ids},
                "status": PositionStatus.CLOSED.value,
                "closed_at": date_filter,
            }
            positions = await Position.find(pos_query).to_list()
        else:
            positions = await Position.find({"user_id": {"$in": user_ids}}).to_list()

        for p in positions:
            net_client_pnl += _realised_inr(p, fallback_usd_inr)

        # Brokerage from trades (wallet.total_brokerage is often 0 because
        # brokerage is tracked per-trade, not as a separate wallet txn)
        bkg_trade_q: dict[str, Any] = {"user_id": {"$in": user_ids}}
        if date_filter:
            bkg_trade_q["executed_at"] = date_filter
        bkg_trades = await Trade.find(bkg_trade_q).to_list()
        for t in bkg_trades:
            net_client_bkg += abs(to_decimal(t.brokerage))

        # Deposits in window
        dep_query: dict[str, Any] = {
            "user_id": {"$in": user_ids},
            "transaction_type": TransactionType.DEPOSIT.value,
        }
        if date_filter:
            dep_query["created_at"] = date_filter

        dep_txns = await WalletTransaction.find(dep_query).to_list()
        for t in dep_txns:
            total_deposits += abs(to_decimal(t.amount))

        # Withdrawals in window
        wd_query: dict[str, Any] = {
            "user_id": {"$in": user_ids},
            "transaction_type": TransactionType.WITHDRAWAL.value,
        }
        if date_filter:
            wd_query["created_at"] = date_filter

        wd_txns = await WalletTransaction.find(wd_query).to_list()
        for t in wd_txns:
            total_withdrawals += abs(to_decimal(t.amount))

    # Settlement outstanding from wallets (always current, not windowed)
    settlement = Decimal("0")
    if user_ids:
        wallets = await Wallet.find({"user_id": {"$in": user_ids}}).to_list()
        for w in wallets:
            settlement += to_decimal(w.settlement_outstanding)

    # Broker view
    broker_view_pnl = -net_client_pnl
    total_of_both = broker_view_pnl + net_client_bkg
    actual_pnl = total_of_both - settlement

    # PNL sharing agreement lookup
    share_pct = Decimal("0")
    agreement_type: str | None = None
    sharing_pnl = Decimal("0")
    sharing_bkg = Decimal("0")

    agreement = await PnlSharingAgreement.find_one({
        "broker_id": entity_id,
        "status": {"$in": [AgreementStatus.ACTIVE.value, AgreementStatus.PAUSED.value]},
    })
    if agreement:
        share_pct = to_decimal(agreement.share_pct)
        agreement_type = agreement.agreement_type.value if hasattr(agreement.agreement_type, "value") else str(agreement.agreement_type)
        share_frac = share_pct / Decimal("100")
        sharing_pnl = quantize_money(broker_view_pnl * share_frac)
        sharing_bkg = quantize_money(net_client_bkg * share_frac)
        if agreement.agreement_type == AgreementType.BROKERAGE_ONLY:
            sharing_pnl = Decimal("0")

    return {
        "net_client_pnl": str(quantize_money(net_client_pnl)),
        "net_client_bkg": str(quantize_money(net_client_bkg)),
        "total_of_both": str(quantize_money(total_of_both)),
        "settlement": str(quantize_money(settlement)),
        "actual_pnl": str(quantize_money(actual_pnl)),
        "sharing_pnl": str(quantize_money(sharing_pnl)),
        "sharing_bkg": str(quantize_money(sharing_bkg)),
        "total_deposits": str(quantize_money(total_deposits)),
        "total_withdrawals": str(quantize_money(total_withdrawals)),
        "share_pct": str(share_pct),
        "agreement_type": agreement_type,
        "client_count": len(user_ids),
    }


def _empty_broker_totals() -> dict[str, Any]:
    z = "0.00"
    return {
        "net_client_pnl": z, "net_client_bkg": z, "total_of_both": z,
        "settlement": z, "actual_pnl": z, "sharing_pnl": z, "sharing_bkg": z,
        "total_deposits": z, "total_withdrawals": z,
        "share_pct": "0", "agreement_type": None, "client_count": 0,
    }


# ── Per-user PNL breakdown within an entity ──────────────────────────

async def get_entity_users(
    entity_id: PydanticObjectId,
    entity_role: str,
    start_utc: datetime | None,
    end_utc: datetime | None,
    page: int = 1,
    page_size: int = 15,
    search: str | None = None,
) -> dict[str, Any]:
    """Paginated per-user PNL breakdown for an entity's client pool."""
    user_query: dict[str, Any] = {
        "role": {"$in": TRADING_ROLES},
        "status": {"$ne": UserStatus.CLOSED.value},
    }
    if entity_role in (UserRole.BROKER.value, "BROKER"):
        user_query["assigned_broker_id"] = entity_id
    elif entity_role in (UserRole.ADMIN.value, "ADMIN"):
        # Include direct users + users under admin's brokers
        broker_ids = [
            b["_id"] async for b in User.get_motor_collection().find(
                {"assigned_admin_id": entity_id, "role": UserRole.BROKER.value},
                {"_id": 1},
            )
        ]
        scope_filter = [{"assigned_admin_id": entity_id}]
        if broker_ids:
            scope_filter.append({"assigned_broker_id": {"$in": broker_ids}})
        user_query["$or"] = scope_filter

    if search and search.strip():
        import re
        escaped = re.escape(search.strip())
        search_filter = [
            {"user_code": {"$regex": escaped, "$options": "i"}},
            {"full_name": {"$regex": escaped, "$options": "i"}},
        ]
        if "$or" in user_query:
            user_query = {"$and": [user_query, {"$or": search_filter}]}
        else:
            user_query["$or"] = search_filter

    total = await User.find(user_query).count()
    total_pages = max(1, (total + page_size - 1) // page_size)
    page = min(page, total_pages)
    skip = (page - 1) * page_size

    users = await User.find(user_query).skip(skip).limit(page_size).to_list()

    if not users:
        return {"items": [], "meta": {"page": page, "page_size": page_size, "total": total, "total_pages": total_pages}}

    date_filter: dict[str, Any] = {}
    if start_utc:
        date_filter["$gte"] = start_utc
    if end_utc:
        date_filter["$lte"] = end_utc
    has_date = bool(date_filter)

    fallback_usd_inr = to_decimal(market_data_service.get_usd_inr_rate())
    items = []

    for u in users:
        # Realized PNL — include ALL positions (open with partial close PNL + closed)
        net_pnl = Decimal("0")
        if has_date:
            # Date-filtered: only closed positions in window
            closed_q: dict[str, Any] = {
                "user_id": u.id,
                "status": PositionStatus.CLOSED.value,
                "closed_at": date_filter,
            }
            for p in await Position.find(closed_q).to_list():
                net_pnl += _realised_inr(p, fallback_usd_inr)
        else:
            # Lifetime: sum realized_pnl from ALL positions (open + closed)
            all_pos = await Position.find({"user_id": u.id}).to_list()
            for p in all_pos:
                net_pnl += _realised_inr(p, fallback_usd_inr)

        # Brokerage from trades
        trade_q: dict[str, Any] = {"user_id": u.id}
        if has_date:
            trade_q["executed_at"] = date_filter
        trades = await Trade.find(trade_q).to_list()
        net_bkg = sum((abs(to_decimal(t.brokerage)) for t in trades), Decimal("0"))

        total_pnl = net_pnl - net_bkg

        # Settlement outstanding (current)
        wallet = await Wallet.find_one({"user_id": u.id})
        settlement_val = to_decimal(wallet.settlement_outstanding) if wallet else Decimal("0")
        pnl_minus_settlement = total_pnl - settlement_val

        items.append({
            "user_id": str(u.id),
            "user_code": u.user_code or "",
            "username": u.full_name or u.user_code or "",
            "net_pnl": str(quantize_money(net_pnl)),
            "net_bkg": str(quantize_money(net_bkg)),
            "total_pnl": str(quantize_money(total_pnl)),
            "settlement": str(quantize_money(settlement_val)),
            "pnl_minus_settlement": str(quantize_money(pnl_minus_settlement)),
        })

    # Sort by total_pnl descending (highest first)
    items.sort(key=lambda x: float(x["total_pnl"]), reverse=True)

    return {
        "items": items,
        "meta": {
            "page": page,
            "page_size": page_size,
            "total": total,
            "total_pages": total_pages,
        },
    }


async def get_all_entity_users(
    entity_id: PydanticObjectId,
    entity_role: str,
    start_utc: datetime | None,
    end_utc: datetime | None,
) -> list[dict[str, Any]]:
    """All users (no pagination) for export."""
    result = await get_entity_users(
        entity_id, entity_role, start_utc, end_utc,
        page=1, page_size=10000,
    )
    return result["items"]


# ── Excel renderers ──────────────────────────────────────────────────

_BOLD = Font(bold=True)
_GREY_FILL = PatternFill("solid", fgColor="EEEEEE")
_RIGHT = Alignment(horizontal="right")
_NUM_FMT = '#,##0.00'


def render_entity_users_excel(
    entity_name: str,
    users_data: list[dict[str, Any]],
    period_label: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "PNL Report"

    ws.append([f"PNL Report — {entity_name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {period_label}"])
    ws.append([])

    headers = ["User ID", "Username", "Total PNL", "Net PNL", "Net BKG", "Settlement", "PNL - Settlement"]
    ws.append(headers)
    row_num = 4
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = _BOLD
        cell.fill = _GREY_FILL

    for u in users_data:
        ws.append([
            u["user_code"],
            u["username"],
            float(u["total_pnl"]),
            float(u["net_pnl"]),
            float(u["net_bkg"]),
            float(u["settlement"]),
            float(u["pnl_minus_settlement"]),
        ])

    for col_idx in range(3, 8):
        for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                cell.number_format = _NUM_FMT
                cell.alignment = _RIGHT

    widths = [12, 20, 15, 15, 15, 15, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


def render_broker_totals_excel(
    totals: dict[str, Any],
    entity_name: str,
    period_label: str,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Broker Summary"

    ws.append([f"Broker Summary — {entity_name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {period_label}"])
    ws.append([])

    rows = [
        ("NET CLIENT PNL", totals["net_client_pnl"]),
        ("NET CLIENT BKG", totals["net_client_bkg"]),
        ("TOTAL OF BOTH", totals["total_of_both"]),
        ("SETTLEMENT", totals["settlement"]),
        ("ACTUAL PNL", totals["actual_pnl"]),
        ("SHARING PNL", totals["sharing_pnl"]),
        ("SHARING BKG", totals["sharing_bkg"]),
        ("TOTAL DEPOSITS", totals["total_deposits"]),
        ("TOTAL WITHDRAWALS", totals["total_withdrawals"]),
    ]
    for label, val in rows:
        ws.append([label, float(val)])

    for row_idx in range(4, 4 + len(rows)):
        ws.cell(row=row_idx, column=1).font = _BOLD
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.number_format = _NUM_FMT
        cell_b.alignment = _RIGHT

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 20

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


def render_single_user_excel(user_data: dict[str, Any], period_label: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "User PNL"

    ws.append([f"PNL Report — {user_data['user_code']} ({user_data['username']})"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {period_label}"])
    ws.append([])

    rows = [
        ("Net PNL", user_data["net_pnl"]),
        ("Net BKG", user_data["net_bkg"]),
        ("Total PNL", user_data["total_pnl"]),
        ("Settlement", user_data["settlement"]),
        ("PNL - Settlement", user_data["pnl_minus_settlement"]),
    ]
    for label, val in rows:
        ws.append([label, float(val)])

    for row_idx in range(4, 4 + len(rows)):
        ws.cell(row=row_idx, column=1).font = _BOLD
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.number_format = _NUM_FMT
        cell_b.alignment = _RIGHT

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data


# ── PDF renderers ────────────────────────────────────────────────────

def _money_fmt(val: str) -> str:
    try:
        f = float(val)
        sign = "+" if f > 0 else ""
        return f"{sign}{f:,.2f}"
    except (ValueError, TypeError):
        return val


def render_entity_users_pdf(
    entity_name: str,
    users_data: list[dict[str, Any]],
    period_label: str,
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title=f"PNL Report — {entity_name}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=14, spaceAfter=6)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements: list = []
    elements.append(Paragraph(f"PNL Report — {entity_name}", h1))
    elements.append(Paragraph(f"Period: {period_label}", meta_style))
    elements.append(Spacer(1, 8))

    headers = ["User ID", "Username", "Total PNL", "Net PNL", "Net BKG", "Settlement", "PNL-Sett."]
    data = [headers]
    for u in users_data:
        data.append([
            u["user_code"], u["username"],
            _money_fmt(u["total_pnl"]), _money_fmt(u["net_pnl"]),
            _money_fmt(u["net_bkg"]), _money_fmt(u["settlement"]),
            _money_fmt(u["pnl_minus_settlement"]),
        ])

    col_widths = [50, 70, 55, 55, 50, 50, 55]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#333333")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (2, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elements.append(tbl)

    doc.build(elements)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes


def render_broker_totals_pdf(
    totals: dict[str, Any],
    entity_name: str,
    period_label: str,
) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=f"Broker Summary — {entity_name}",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, spaceAfter=8)
    meta_style = ParagraphStyle("meta", parent=styles["Normal"], fontSize=9, textColor=colors.grey)

    elements: list = []
    elements.append(Paragraph(f"Broker Summary — {entity_name}", h1))
    elements.append(Paragraph(f"Period: {period_label}", meta_style))
    elements.append(Spacer(1, 10))

    rows = [
        ["NET CLIENT PNL", _money_fmt(totals["net_client_pnl"])],
        ["NET CLIENT BKG", _money_fmt(totals["net_client_bkg"])],
        ["TOTAL OF BOTH", _money_fmt(totals["total_of_both"])],
        ["− SETTLEMENT", _money_fmt(totals["settlement"])],
        ["= ACTUAL PNL", _money_fmt(totals["actual_pnl"])],
        ["SHARING PNL", _money_fmt(totals["sharing_pnl"])],
        ["SHARING BKG", _money_fmt(totals["sharing_bkg"])],
        ["TOTAL DEPOSITS", _money_fmt(totals["total_deposits"])],
        ["TOTAL WITHDRAWALS", _money_fmt(totals["total_withdrawals"])],
    ]
    tbl = Table(rows, colWidths=[70 * mm, 60 * mm])
    tbl.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#e8f5e9")),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)

    doc.build(elements)
    pdf_bytes = buf.getvalue()
    buf.close()
    return pdf_bytes
